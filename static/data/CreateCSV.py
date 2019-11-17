
# coding: utf-8

# In[1]:


import pandas as pd
file = 'HBW-BearLife_Checklist_Version_3.xlsx'


# In[2]:


data = pd.read_excel(file)
data.columns = data.iloc[0]
data = data.drop(0)


# In[3]:


import openpyxl

wb = openpyxl.load_workbook(file)
ws = wb.active


# In[13]:


hyperlinks = []
for row in range(3,ws.max_row+1):
    hyperlinks.append(ws.cell(row=row, column=5)
                      .hyperlink.target.split('/')[-1])


# In[14]:


data['hyperlink'] = hyperlinks


# In[15]:


names = [
    'Order','Family name','Family','CommonName','ScientificName',
    'Authority','TaxonomicTreatment', 'IUCNRedListCategory','Synonyms', 
    'AlternativeCommonNames', 'TaxonomicNotes','TaxonomicSource', 'SISRecID','SpcRecID', 'Hyperlink'
]
data.columns = names
data.head()


# In[17]:


data.to_csv("Osos.csv", index=False)
data.tail()

